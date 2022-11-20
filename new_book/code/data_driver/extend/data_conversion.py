import json
import yaml
from extend.parameterized_extend import check_data

import csv
import codecs
from itertools import islice

from openpyxl import load_workbook


import os
from extend.parameterized_extend import data


def file_data(file: str, line: int = 1, sheet: str = "Sheet1", key: str = None):
    """
    支持文件转参数化.

    :param file: 文件名
    :param line:  Excel/CSV 文件指定第几行开始读取
    :param sheet: Excel文件标签页名
    :param key: YAML/JSON文件指定key
    """
    if file is None:
        raise FileExistsError("File name does not exist.")

    if os.path.isfile(file) is False:
        raise FileExistsError(f"No '{file}' data file found.")

    suffix = file.split(".")[-1]
    if suffix == "csv":
        data_list = csv_to_list(file, line=line)
    elif suffix == "xlsx":
        data_list = excel_to_list(file, sheet=sheet, line=line)
    elif suffix == "json":
        data_list = json_to_list(file, key=key)
    elif suffix == "yaml":
        data_list = yaml_to_list(file, key=key)
    else:
        raise FileExistsError(f"Your file is not supported: {file}")

    return data(data_list)


def json_to_list(file: str = None, key: str = None) -> list:
    """
    将JSON文件数据转换为list
    :param file: JSON文件地址
    :param key: 字典的key
    return: list
    """
    if file is None:
        raise FileExistsError("Please specify the JSON file to convert.")

    if key is None:
        with open(file, "r", encoding="utf-8") as json_file:
            data = json.load(json_file)
            list_data = check_data(data)
    else:
        with open(file, "r", encoding="utf-8") as json_file:
            try:
                data = json.load(json_file)[key]
                list_data = check_data(data)
            except KeyError:
                raise ValueError(f"Check the test data, no '{key}'.")

    return list_data


def yaml_to_list(file: str = None, key: str = None) -> list:
    """
    将YAML文件数据转换为list
    :param file: YAML文件路径
    :param key: 字典的key
    return: list
    """
    if file is None:
        raise FileExistsError("Please specify the YAML file to convert.")

    if key is None:
        with open(file, "r", encoding="utf-8") as yaml_file:
            data = yaml.load(yaml_file, Loader=yaml.FullLoader)
            list_data = check_data(data)
    else:
        with open(file, "r", encoding="utf-8") as yaml_file:
            try:
                data = yaml.load(yaml_file, Loader=yaml.FullLoader)[key]
                list_data = check_data(data)
            except KeyError as exc:
                raise ValueError(f"Check the YAML test data, no '{key}'") from exc

    return list_data


def csv_to_list(file: str = None, line: int = 1) -> list:
    """
    将CSV文件数据转换为list
    :param file: CSV文件路径
    :param line: 从第几行开始读取
    :return: list
    """
    if file is None:
        raise FileExistsError("Please specify the CSV file to convert.")

    table_data = []
    with codecs.open(file, 'r', encoding='utf_8_sig') as csv_file:
        csv_data = csv.reader(csv_file)
        for i in islice(csv_data, line - 1, None):
            table_data.append(i)

    return table_data


def excel_to_list(file: str = None, sheet: str = "Sheet1", line: int = 1) -> list:
    """
    将Excel文件数据转换为list
    :param file: Excel文件路径
    :param sheet: 文件标签页的名字, 默认：Sheet1
    :param line: 从第几行开始读取
    :return: list data
    """
    if file is None:
        raise FileExistsError("Please specify the Excel file to convert.")

    excel_table = load_workbook(file)
    sheet = excel_table[sheet]

    table_data = []
    for i in sheet.iter_rows(line, sheet.max_row):
        line_data = []
        for field in i:
            line_data.append(field.value)
        table_data.append(line_data)

    return table_data
